#coding:utf-8

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import coordinate_from_string, column_index_from_string
import sys
from datetime import datetime, timedelta, date

#https://www.jianshu.com/p/ce2ba7caa414

reload(sys)
sys.setdefaultencoding('utf-8')
dt = date(2018, 3, 1)

#读取分表
def reloadExcel(fileExcel):

    #读到文件
    wb = load_workbook(filename = fileExcel, data_only=True)

    #读取汇总文件并删除
    totalDataDictionary = reloadTotalExcel(wb)

    #根据名称获取工作表
    #sheet_ranges = wb['1']

    #获取单元格值
    #cell_value = sheet_ranges['D18'].value

    dataDictionary = {}
    #循环所有工作表
    # for sheet in wb:
    #     print(sheet.title)

    #获取单元格区域
    # cell_range = sheet_ranges['B4':'U4']
    # for row in cell_range:
    #     for cell in row:
    #         print(cell.value)

    for sheet in wb:
        # for row in sheet['A4':'U4']:
        for row in sheet.iter_rows(min_row=4, max_col=14, max_row=47):
            dataSheetData = []
            dataRowList = []
            for cell in row:
                dataRowList.append(cell.value)
            if dataDictionary.has_key(sheet.cell(row=cell.row, column=1).value):
                dataDictionary[sheet.cell(row=cell.row, column=1).value].append(dataRowList)
            else:
                dataSheetData.append(dataRowList)
                dataDictionary[sheet.cell(row=cell.row, column=1).value] = dataSheetData


    # print(dataDictionary)            

    copyNewExcel(dataDictionary, totalDataDictionary)

def reloadTotalExcel(wb):

    #wb = load_workbook(filename = fileExcel, data_only=True)
    sheet = wb.worksheets[0]

    dataDictionary = {}

    # for row in sheet['A4':'U4']:
    for row in sheet.iter_rows(min_row=2, max_col=44, max_row=45):
        dataSheetData = []
        dataRowList = []
        for cell in row:
            dataRowList.append(cell.value)
        if dataDictionary.has_key(sheet.cell(row=cell.row, column=1).value):
            dataDictionary[sheet.cell(row=cell.row, column=1).value].append(dataRowList)
        else:
            dataSheetData.append(dataRowList)
            dataDictionary[sheet.cell(row=cell.row, column=1).value] = dataSheetData

    wb.remove(sheet)
    return dataDictionary;

#为模板赋值
def copyNewExcel(dataDictionary, totalDataDictionary):

    #----------------------------华丽的分隔线---------------------------------------
    #读到模板文件
    wb_template = load_workbook(filename = 'HBH-个人.xlsx', data_only=True)

    #获取工作表
    template_sheet_range = wb_template['Sheet1']

    #设置单元格值
    # template_sheet_range['D8'] = '这是一个测试'

    #设置工作表名称
    # template_sheet_range.title = 'test'
    #复制新的工作表
    # copy_sheet_range = wb_template.copy_worksheet(template_sheet_range)
    # copy_sheet_range.title = 'copytest'

    #添加新的工作表
    #ws1 = wb_template.create_sheet('Mysheet')

    # setSheeBoder(template_sheet_range, copy_sheet_range);


    for k, v in dataDictionary.items():
        copy_sheet_range = wb_template.copy_worksheet(template_sheet_range)
        copy_sheet_range.title = str(k)
        setSheeBoder(template_sheet_range, copy_sheet_range)
        copy_sheet_range['A3'] = v[0][1]
        copy_sheet_range['C3'] = str('')
        copy_sheet_range['E3'] = str('')
        copy_sheet_range['K3'] = str('')
        #银卡
        copy_sheet_range['F4'] = 0 if totalDataDictionary[k][0][36] == None else totalDataDictionary[k][0][36]
        #金卡
        copy_sheet_range['G4'] = 0 if totalDataDictionary[k][0][37] == None else totalDataDictionary[k][0][37]
        #明珠
        copy_sheet_range['H4'] = 0 if totalDataDictionary[k][0][39] == None else totalDataDictionary[k][0][39]
        #明珠
        copy_sheet_range['I4'] = 0 if totalDataDictionary[k][0][41] == None else totalDataDictionary[k][0][41]
        #推荐
        copy_sheet_range['J4'] = 0 if totalDataDictionary[k][0][43] == None else totalDataDictionary[k][0][43]
        #合计
        copy_sheet_range['F37'] = 0 if totalDataDictionary[k][0][2] == None else totalDataDictionary[k][0][2]

        for index, row in enumerate(v):
            for column in range(2, 7):
                copy_sheet_range.cell(row = index + 6, column = column).value = row[column + 7]
            #日期
            copy_sheet_range.cell(row = index + 6, column = 1).value = '{}{}{}'.format(monthFormat((dt + timedelta(days=index)).strftime('%m')), '/', monthFormat((dt + timedelta(days=index)).strftime('%d')))

    wb_template.remove(wb_template['Sheet1']);

    saveExcel(wb_template);





#合并单元格样式
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def setSheeBoder(template_sheet_range, copy_sheet_range):

    thin = Side(border_style="thin", color="000000")
    single = Side(border_style="thin", color="000000")

    border = Border(top=single, left=thin, right=thin, bottom=single)

    #------------------------------------华丽的分隔线------------------------------------------------                

    style_range(template_sheet_range, 'A1:B1', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'A3:B4', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'C2:D2', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'C3:D3', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'F2:J3', border=border, fill=None, font=None, alignment=None)
    # style_range(template_sheet_range, 'A4:J5', border=border, fill=None, font=None, alignment=None)
    
    style_range(template_sheet_range, 'E3:E4', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'K3:K4', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'A37:E37', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'H37:K37', border=border, fill=None, font=None, alignment=None)
    style_range(template_sheet_range, 'F37:G37', border=border, fill=None, font=None, alignment=None)

    style_range(copy_sheet_range, 'A1:B1', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'A3:B4', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'C2:D2', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'C3:D3', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'F2:J3', border=border, fill=None, font=None, alignment=None)
    # style_range(copy_sheet_range, 'A4:J5', border=border, fill=None, font=None, alignment=None)
    
    style_range(copy_sheet_range, 'E3:E4', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'K3:K4', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'A37:E37', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'H37:K37', border=border, fill=None, font=None, alignment=None)
    style_range(copy_sheet_range, 'F37:G37', border=border, fill=None, font=None, alignment=None)

    for row in range(6, 37):
        style_range(template_sheet_range, ('F' + str(row)) + ':' + ('G' + str(row)), border=border, fill=None, font=None, alignment=None)
        style_range(template_sheet_range, ('I' + str(row)) + ':' + ('J' + str(row)), border=border, fill=None, font=None, alignment=None)

        style_range(copy_sheet_range, ('F' + str(row)) + ':' + ('G' + str(row)), border=border, fill=None, font=None, alignment=None)
        style_range(copy_sheet_range, ('I' + str(row)) + ':' + ('J' + str(row)), border=border, fill=None, font=None, alignment=None)


def saveExcel(wb_template):
    #保存工作簿 template:是否做为模板
    wb_template.template = False
    wb_template.save('HBH.xlsx')

def weekday(w):
    wd = {'0': '日', '1': '一', '2': '二', '3': '三', '4': '四', '5': '五', '6': '六'}
    return wd[w]

def monthFormat(m):
    if m.startswith('0', 0, 1):
        return m[1:]
    else:
        return m

def main():
    print('开始处理HBH-EXCEL----------')
    reloadExcel('HBH-3月.xlsx')

main()
